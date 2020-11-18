from googletrans import Translator
import openpyxl
import time
import math

# Intro Message
print('This is auto translation program which takes excel as input and translates entire file based on user input')
print('This is created by Shridhar Sahu. Please contact him in case of any queries\n')

# Reading Data from Excel, Setting up Translation Object and Dictionaries
print('Reading user inputs ...')
time.sleep(1)
workbook = openpyxl.load_workbook('Input_Translation.xlsx')
Data = workbook['Data']
InputSelect = workbook['Selection']
TranslationSite = InputSelect.cell(row=1, column=2).value
SourceLang = InputSelect.cell(row=2, column=2).value
DestLang = InputSelect.cell(row=3, column=2).value
GoogleTransLang = {'Auto Detect': 'auto', 'Afrikaans': 'af', 'Albanian': 'sq', 'Amharic': 'am', 'Arabic': 'ar', 'Armenian': 'hy', 'Azerbaijani': 'az', 'Basque': 'eu', 'Belarusian': 'be', 'Bengali': 'bn', 'Bosnian': 'bs', 'Bulgarian': 'bg', 'Catalan': 'ca', 'Cebuano': 'ceb', 'Chichewa': 'ny', 'Chinese (Simplified)': 'zh-cn', 'Chinese (Traditional)': 'zh-tw', 'Corsican': 'co', 'Croatian': 'hr', 'Czech': 'cs', 'Danish': 'da', 'Dutch': 'nl', 'English': 'en', 'Esperanto': 'eo', 'Estonian': 'et', 'Filipino': 'tl', 'Finnish': 'fi', 'French': 'fr', 'Frisian': 'fy', 'Galician': 'gl', 'Georgian': 'ka', 'German': 'de', 'Greek': 'el', 'Gujarati': 'gu', 'Haitian Creole': 'ht', 'Hausa': 'ha', 'Hawaiian': 'haw', 'Hebrew': 'iw', 'Hebrew': 'he', 'Hindi': 'hi', 'Hmong': 'hmn', 'Hungarian': 'hu', 'Icelandic': 'is', 'Igbo': 'ig', 'Indonesian': 'id', 'Irish': 'ga', 'Italian': 'it', 'Japanese': 'ja', 'Javanese': 'jw', 'Kannada': 'kn', 'Kazakh': 'kk', 'Khmer': 'km', 'Korean': 'ko', 'Kurdish (Kurmanji)': 'ku', 'Kyrgyz': 'ky', 'Lao': 'lo', 'Latin': 'la', 'Latvian': 'lv', 'Lithuanian': 'lt', 'Luxembourgish': 'lb', 'Macedonian': 'mk', 'Malagasy': 'mg', 'Malay': 'ms', 'Malayalam': 'ml', 'Maltese': 'mt', 'Maori': 'mi', 'Marathi': 'mr', 'Mongolian': 'mn', 'Myanmar (Burmese)': 'my', 'Nepali': 'ne', 'Norwegian': 'no', 'Odia': 'or', 'Pashto': 'ps', 'Persian': 'fa', 'Polish': 'pl', 'Portuguese': 'pt', 'Punjabi': 'pa', 'Romanian': 'ro', 'Russian': 'ru', 'Samoan': 'sm', 'Scots Gaelic': 'gd', 'Serbian': 'sr', 'Sesotho': 'st', 'Shona': 'sn', 'Sindhi': 'sd', 'Sinhala': 'si', 'Slovak': 'sk', 'Slovenian': 'sl', 'Somali': 'so', 'Spanish': 'es', 'Sundanese': 'su', 'Swahili': 'sw', 'Swedish': 'sv', 'Tajik': 'tg', 'Tamil': 'ta', 'Telugu': 'te', 'Thai': 'th', 'Turkish': 'tr', 'Ukrainian': 'uk', 'Urdu': 'ur', 'Uyghur': 'ug', 'Uzbek': 'uz', 'Vietnamese': 'vi', 'Welsh': 'cy', 'Xhosa': 'xh', 'Yiddish': 'yi', 'Yoruba': 'yo', 'Zulu': 'zu'}
SourceCode = GoogleTransLang[SourceLang]
DestCode = GoogleTransLang[DestLang]
translator = Translator()
TranslationDict = {}
print('Source Language is %s' %(SourceLang))
print('Destination Language is %s' %(DestLang))
print('We are using %s translation services\n' %(TranslationSite))
time.sleep(1)

# Reading through entire data set and storing data in Dict
print('Reading excel data for translation ...')
for r in range(2, Data.max_row+1):
    if Data.cell(row=r, column=1).value in [None, '']:
        break
    for c in range(2, Data.max_column+1):
        if Data.cell(row=1, column=c) in [None, '']:
            break
        if Data.cell(row=r, column=c).value not in [None, '']:
            Text = str(Data.cell(row=r, column=c).value).strip()
            TranslationDict[Text] = None

print('There are %s unique items for translation ...' % (len(list(TranslationDict.keys()))))

# Translating the entire dictionary
try:
    TranslationList = list(TranslationDict.keys())
    runcounter = math.ceil(len(TranslationList) / 100)
    for i in range(runcounter):
        if i == 0:
            start, end = 0, 101
        elif i == (runcounter - 1):
            start, end = (i*100) + 1, len(TranslationList)
        else:
            start, end = (i*100) + 1, (i*100) + 101
        PartialList = TranslationList[start:end]
        translations = translator.translate(PartialList, src=SourceCode, dest=DestCode)
        for translation in translations:
            TranslationDict[translation.origin] = translation.text
        if i == (runcounter - 1):
            print('%s unique items have been translated' % end)
        else:
            time.sleep(50)
            print('%s unique items have been translated ...' % (end-1))
except:
    print('Oops something failed during translations')

# Store the dictionary value back in cells
print('\nTranslation is done. Storing data in output files ...')
for r in range(2, Data.max_row+1):
    if Data.cell(row=r, column=1).value in [None, '']:
        break
    for c in range(2, Data.max_column+1):
        if Data.cell(row=1, column=c) in [None, '']:
            break
        if Data.cell(row=r, column=c).value not in [None, '']:
            Text = str(Data.cell(row=r, column=c).value).strip()
            Data.cell(row=r, column=c).value = TranslationDict.get(Text, '')

OutputFile = 'Output_' + SourceLang + '_to_' + DestLang + '_' + time.strftime("%Y%m%d_%H%M%S") + '.xlsx'
workbook.remove(InputSelect)
workbook.save(OutputFile)
print('Output is saved as %s' % OutputFile)
